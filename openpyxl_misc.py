import os
import openpyxl

dossier     = '.'
fichier     = 'a.xlsx'
feuille     = 'Feuil1'

wb = openpyxl.load_workbook(fichier, read_only=False)
ws = wb[feuille]
ws['A1'] = "coucou"
wb.save(fichier)
wb.close()



'''
# Lister tous les fichiers XLSX du répertoire et toutes les feuilles qu'ils contiennent :
files = os.listdir(dossier)
files = [f for f in files if os.path.isfile(f)]  # exclure les répertoires
files = [f for f in files if f.split(".")[-1].lower() == "xlsx"]  # ne retenir que les fichiers XLSX
print("\nFichiers XLSX :", files)

for f in files:
    print("\nClasseur :", f)
    wb = openpyxl.load_workbook(f, read_only=False)
    sheets = wb.sheetnames
    for s in sheets:
        print("\tFeuille :", s)
        # traitement...
    wb.save(f)
    wb.close()

input("\n\nAppuyez sur Entrée pour quitter")
'''

