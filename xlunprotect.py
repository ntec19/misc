#!/usr/bin/env python3

import sys

try:
    import openpyxl
except ImportError:
    print("Bibliothèque 'openpyxl' introuvable !")
    exit(1)

if len(sys.argv) != 2:
    print(f"Usage : {sys.argv[0]} <fichierExcelADéprotéger>")
    exit(2)

fichier = sys.argv[1]

# retirer la protection du classeur
wb = openpyxl.load_workbook(fichier)
try:
    wb.security.lockStructure = False
except:
    print("Classeur non verrouillé")
wb.save(fichier)
wb.close()

# retirer la protection de toutes les feuilles
wb = openpyxl.load_workbook(fichier)
liste_feuilles = wb.sheetnames
for feuille in liste_feuilles:
    ws = wb[feuille]
    ws.protection.disable()
    print(f"Protection retirée sur la feuille {feuille}")
wb.save(fichier)
wb.close()

print("terminé !\n")
